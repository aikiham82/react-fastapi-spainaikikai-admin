"""Load and normalize Excel rows for sync."""

from __future__ import annotations

import datetime
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .constants import SHEET_FEES, SHEET_INSURANCES, SHEET_MEMBERS, TAIO_REMAP
from .normalizers import norm_dni


@dataclass
class ExcelMemberRow:
    num_socio: str
    first_name: str
    last1: str
    last2: str
    dni_raw: str
    email: str
    phone: str
    birth_date: datetime.date | None
    address: str
    city: str
    province: str
    postal_code: str
    country: str
    club_id: str
    club_name: str


@dataclass
class ExcelFeeRow:
    num_socio: str
    grade_level: int | None  # "Nivel" column
    grade_type: str  # "dan" | "kyu" | "kyu_infantil" | ""
    instructor: str
    cuota_anual: float
    seguro_accidentes: float
    seguro_rc_flag: bool
    send_date: datetime.datetime | None


@dataclass
class ExcelInsuranceRow:
    num_socio_hint: str
    first_name: str
    last1: str
    last2: str
    dni_raw: str
    club_name: str
    tipo_seguro: str  # e.g., "seguro_accidentes - 2026"
    start_date: datetime.datetime | None
    end_date: datetime.datetime | None


def _s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num_socio(v) -> str:
    """Normalize num_socio across int/float/str Excel cell types."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _to_date(v) -> datetime.date | None:
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _apply_taio_remap(club_name: str, club_id: str) -> str:
    return TAIO_REMAP.get((club_name.upper(), club_id), club_id)


def load_members(path: Path) -> list[ExcelMemberRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_MEMBERS]
    rows: list[ExcelMemberRow] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        club_name = _s(row[16])
        club_id = _s(row[13])
        club_id = _apply_taio_remap(club_name, club_id)
        rows.append(
            ExcelMemberRow(
                num_socio=_num_socio(row[0]),
                first_name=_s(row[1]),
                last1=_s(row[2]),
                last2=_s(row[3]),
                dni_raw=_s(row[4]),
                email=_s(row[5]),
                phone=_s(row[6]),
                birth_date=_to_date(row[7]),
                address=_s(row[8]),
                city=_s(row[9]),
                province=_s(row[10]),
                postal_code=_s(row[11]),
                country=_s(row[12]) or "Spain",
                club_id=club_id,
                club_name=club_name,
            )
        )
    return rows


def load_fees(path: Path) -> dict[str, ExcelFeeRow]:
    """Returns {num_socio: ExcelFeeRow}."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_FEES]
    rows: dict[str, ExcelFeeRow] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        num = _num_socio(row[0])
        nivel = row[4] if isinstance(row[4], (int, float)) else None
        grade_type = _s(row[5]).lower()
        cuota = float(row[7]) if isinstance(row[7], (int, float)) else 0.0
        seg_acc = float(row[8]) if isinstance(row[8], (int, float)) else 0.0
        rc_flag = _s(row[9]).upper() == "RC"
        send_date = row[20] if isinstance(row[20], datetime.datetime) else None
        rows[num] = ExcelFeeRow(
            num_socio=num,
            grade_level=int(nivel) if nivel is not None else None,
            grade_type=grade_type,
            instructor=_s(row[6]),
            cuota_anual=cuota,
            seguro_accidentes=seg_acc,
            seguro_rc_flag=rc_flag,
            send_date=send_date,
        )
    return rows


def load_insurances(path: Path) -> list[ExcelInsuranceRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_INSURANCES]
    rows: list[ExcelInsuranceRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None and row[4] is None:
            continue
        rows.append(
            ExcelInsuranceRow(
                num_socio_hint=_num_socio(row[0]),
                first_name=_s(row[1]),
                last1=_s(row[2]),
                last2=_s(row[3]),
                dni_raw=_s(row[4]),
                club_name=_s(row[5]),
                tipo_seguro=_s(row[6]),
                start_date=row[10] if isinstance(row[10], datetime.datetime) else None,
                end_date=row[11] if isinstance(row[11], datetime.datetime) else None,
            )
        )
    return rows


def member_has_matchable_identity(r: ExcelMemberRow) -> bool:
    """Returns True if we have anything to match on (DNI or club+name)."""
    return bool(norm_dni(r.dni_raw)) or (bool(r.club_id) and bool(r.first_name))
