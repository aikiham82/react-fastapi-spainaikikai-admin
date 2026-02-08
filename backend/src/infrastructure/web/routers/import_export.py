"""Import/Export routes."""

from typing import Optional
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.infrastructure.web.dto.import_export_dto import (
    ImportMembersRequest,
    ImportMembersResponse,
    ImportLicensesRequest,
    ImportInsurancesRequest,
    ImportPaymentsRequest
)
from src.infrastructure.web.dependencies import (
    get_all_members_use_case,
    get_create_member_use_case,
    get_update_member_use_case,
    get_all_licenses_use_case,
    get_all_insurances_use_case,
    get_create_license_use_case,
    get_update_license_use_case,
    get_create_insurance_use_case,
    get_update_insurance_use_case,
    get_member_repository,
    get_license_repository,
    get_insurance_repository,
    get_member_payment_repository,
    get_club_repository
)
from src.infrastructure.web.dependencies import get_auth_context
from src.infrastructure.web.authorization import AuthContext, get_club_filter_ctx
from src.domain.entities.license import LicenseStatus, TechnicalGrade, InstructorCategory, AgeCategory
from src.domain.entities.insurance import InsuranceType, InsuranceStatus
from src.domain.entities.member_payment import MemberPaymentType, MemberPaymentStatus, MemberPayment

router = APIRouter(prefix="/import-export", tags=["import-export"])


def _parse_date(value) -> Optional[datetime]:
    """Parse a date value from Excel data, supporting multiple formats."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


@router.post("/members/import", response_model=ImportMembersResponse)
async def import_members(
    request: ImportMembersRequest,
    create_member_use_case=Depends(get_create_member_use_case),
    update_member_use_case=Depends(get_update_member_use_case),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Import members from Excel data. Supports 'create' and 'upsert' modes."""
    imported = 0
    updated = 0
    failed = 0
    errors = []
    is_upsert = request.mode == "upsert"

    for idx, row in enumerate(request.members):
        try:
            # Map Excel column names to entity fields
            first_name = row.get('first_name') or row.get('Nombre') or row.get('nombre') or ''
            email = row.get('email') or row.get('Email') or row.get('EMAIL') or ''
            last_name = row.get('last_name') or row.get('Apellidos') or row.get('apellidos') or ''
            dni = row.get('dni') or row.get('DNI') or row.get('Dni') or ''
            phone = row.get('phone') or row.get('Teléfono') or row.get('telefono') or ''
            address = row.get('address') or row.get('Dirección') or row.get('direccion') or ''
            city = row.get('city') or row.get('Ciudad') or row.get('ciudad') or ''
            province = row.get('province') or row.get('Provincia') or row.get('provincia') or ''
            postal_code = row.get('postal_code') or row.get('Código Postal') or row.get('codigo_postal') or ''
            country = row.get('country') or row.get('País') or row.get('pais') or 'España'
            club_id = row.get('club_id') or row.get('Club ID') or None
            birth_date_str = row.get('birth_date') or row.get('Fecha Nacimiento') or row.get('fecha_nacimiento') or None

            birth_date = None
            if birth_date_str:
                try:
                    if isinstance(birth_date_str, datetime):
                        birth_date = birth_date_str
                    elif isinstance(birth_date_str, str):
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                birth_date = datetime.strptime(birth_date_str, fmt)
                                break
                            except ValueError:
                                continue
                except Exception:
                    pass

            # In upsert mode, try to find existing member by ID, DNI or email
            if is_upsert:
                existing = None
                member_id = row.get('id') or row.get('ID') or row.get('Id') or ''
                if member_id:
                    existing = await member_repo.find_by_id(str(member_id))
                if not existing and dni:
                    existing = await member_repo.find_by_dni(dni)
                if not existing and email:
                    existing = await member_repo.find_by_email(email)

                if existing:
                    update_fields = {}
                    if first_name:
                        update_fields['first_name'] = first_name
                    if last_name:
                        update_fields['last_name'] = last_name
                    if dni:
                        update_fields['dni'] = dni
                    if email:
                        update_fields['email'] = email
                    if phone:
                        update_fields['phone'] = phone
                    if address:
                        update_fields['address'] = address
                    if city:
                        update_fields['city'] = city
                    if province:
                        update_fields['province'] = province
                    if postal_code:
                        update_fields['postal_code'] = postal_code
                    if country:
                        update_fields['country'] = country
                    if club_id:
                        update_fields['club_id'] = club_id
                    if birth_date:
                        update_fields['birth_date'] = birth_date

                    if update_fields:
                        await update_member_use_case.execute(
                            member_id=existing.id,
                            **update_fields
                        )
                    updated += 1
                    continue

            # For create mode, name and email are required
            if not first_name or not email:
                errors.append(f"Fila {idx + 1}: Nombre y email son obligatorios")
                failed += 1
                continue

            await create_member_use_case.execute(
                first_name=first_name,
                last_name=last_name,
                dni=dni,
                email=email,
                phone=phone,
                address=address,
                city=city,
                province=province,
                postal_code=postal_code,
                country=country,
                club_id=club_id,
                birth_date=birth_date
            )
            imported += 1

        except Exception as e:
            failed += 1
            errors.append(f"Fila {idx + 1}: {str(e)}")

    return ImportMembersResponse(
        success=failed == 0,
        imported=imported,
        updated=updated,
        failed=failed,
        errors=errors
    )


@router.get("/members/export")
async def export_members(
    club_id: Optional[str] = Query(None),
    get_all_use_case=Depends(get_all_members_use_case),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Export members to Excel file. No limit — exports all matching members."""
    # Enforce club isolation: club users can only export their own club's members
    effective_club_id = get_club_filter_ctx(ctx)
    if effective_club_id is not None:
        members = await get_all_use_case.execute(0, effective_club_id)
    elif club_id:
        members = await get_all_use_case.execute(0, club_id)
    else:
        members = await get_all_use_case.execute(0, None)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Miembros"

    # Define headers
    headers = [
        "ID", "Nombre", "Apellidos", "DNI", "Email", "Teléfono",
        "Fecha Nacimiento", "Dirección", "Ciudad", "Provincia",
        "Código Postal", "País", "Club ID", "Estado", "Fecha Creación"
    ]

    # Style for header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, member in enumerate(members, 2):
        ws.cell(row=row_idx, column=1, value=member.id)
        ws.cell(row=row_idx, column=2, value=member.first_name)
        ws.cell(row=row_idx, column=3, value=member.last_name or '')
        ws.cell(row=row_idx, column=4, value='' if not member.dni or member.dni == 'null' else member.dni)
        ws.cell(row=row_idx, column=5, value=member.email)
        ws.cell(row=row_idx, column=6, value=member.phone or '')
        ws.cell(row=row_idx, column=7, value=member.birth_date.strftime('%d/%m/%Y') if member.birth_date else '')
        ws.cell(row=row_idx, column=8, value=member.address or '')
        ws.cell(row=row_idx, column=9, value=member.city or '')
        ws.cell(row=row_idx, column=10, value=member.province or '')
        ws.cell(row=row_idx, column=11, value=member.postal_code or '')
        ws.cell(row=row_idx, column=12, value=member.country or '')
        ws.cell(row=row_idx, column=13, value=member.club_id or '')
        ws.cell(row=row_idx, column=14, value=member.status.value if member.status else '')
        ws.cell(row=row_idx, column=15, value=member.created_at.strftime('%d/%m/%Y %H:%M') if member.created_at else '')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    filename = f"miembros_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/licenses/export")
async def export_licenses(
    club_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    technical_grade: Optional[str] = Query(None),
    age_category: Optional[str] = Query(None),
    get_all_use_case=Depends(get_all_licenses_use_case),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Export licenses to Excel file. No limit — exports all. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden exportar licencias"
        )

    licenses = await get_all_use_case.execute(limit=0, club_id=club_id)

    # Post-filter by status, technical_grade, age_category
    if status:
        try:
            status_enum = LicenseStatus(status)
            licenses = [lic for lic in licenses if lic.status == status_enum]
        except ValueError:
            pass

    if technical_grade:
        try:
            grade_enum = TechnicalGrade(technical_grade)
            licenses = [lic for lic in licenses if lic.technical_grade == grade_enum]
        except ValueError:
            pass

    if age_category:
        try:
            age_enum = AgeCategory(age_category)
            licenses = [lic for lic in licenses if lic.age_category == age_enum]
        except ValueError:
            pass

    # Batch member lookup
    member_ids = list(set([lic.member_id for lic in licenses if lic.member_id]))
    member_map = {}
    for mid in member_ids:
        member = await member_repo.find_by_id(mid)
        if member:
            member_map[mid] = member

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licencias"

    headers = [
        "Nº Licencia", "Nombre", "Apellidos", "DNI", "Club",
        "Grado Técnico", "Cat. Instructor", "Cat. Edad",
        "Estado", "Fecha Emisión", "Fecha Expiración", "Renovada"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, lic in enumerate(licenses, 2):
        member = member_map.get(lic.member_id)
        ws.cell(row=row_idx, column=1, value=lic.license_number)
        ws.cell(row=row_idx, column=2, value=member.first_name if member else '')
        ws.cell(row=row_idx, column=3, value=member.last_name if member else '')
        member_dni = (member.dni if member else '') or ''
        ws.cell(row=row_idx, column=4, value='' if member_dni == 'null' else member_dni)
        ws.cell(row=row_idx, column=5, value=member.club_id if member else '')
        ws.cell(row=row_idx, column=6, value=lic.technical_grade.value if lic.technical_grade else '')
        ws.cell(row=row_idx, column=7, value=lic.instructor_category.value if lic.instructor_category else '')
        ws.cell(row=row_idx, column=8, value=lic.age_category.value if lic.age_category else '')
        ws.cell(row=row_idx, column=9, value=lic.status.value if lic.status else '')
        ws.cell(row=row_idx, column=10, value=lic.issue_date.strftime('%d/%m/%Y') if lic.issue_date else '')
        ws.cell(row=row_idx, column=11, value=lic.expiration_date.strftime('%d/%m/%Y') if lic.expiration_date else '')
        ws.cell(row=row_idx, column=12, value='Sí' if lic.is_renewed else 'No')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"licencias_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/insurances/export")
async def export_insurances(
    club_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    insurance_type: Optional[str] = Query(None),
    get_all_use_case=Depends(get_all_insurances_use_case),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Export insurances to Excel file. No limit — exports all. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden exportar seguros"
        )

    insurances = await get_all_use_case.execute(limit=0, club_id=club_id)

    # Post-filter by status, insurance_type
    if status:
        try:
            status_enum = InsuranceStatus(status)
            insurances = [ins for ins in insurances if ins.status == status_enum]
        except ValueError:
            pass

    if insurance_type:
        try:
            type_enum = InsuranceType(insurance_type)
            insurances = [ins for ins in insurances if ins.insurance_type == type_enum]
        except ValueError:
            pass

    # Batch member lookup
    member_ids = list(set([ins.member_id for ins in insurances if ins.member_id]))
    member_map = {}
    for mid in member_ids:
        member = await member_repo.find_by_id(mid)
        if member:
            member_map[mid] = member

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguros"

    headers = [
        "Nº Póliza", "Nombre", "Apellidos", "DNI", "Club",
        "Tipo Seguro", "Compañía", "Cobertura",
        "Estado", "Fecha Inicio", "Fecha Fin"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ins in enumerate(insurances, 2):
        member = member_map.get(ins.member_id)
        ws.cell(row=row_idx, column=1, value=ins.policy_number)
        ws.cell(row=row_idx, column=2, value=member.first_name if member else '')
        ws.cell(row=row_idx, column=3, value=member.last_name if member else '')
        member_dni = (member.dni if member else '') or ''
        ws.cell(row=row_idx, column=4, value='' if member_dni == 'null' else member_dni)
        ws.cell(row=row_idx, column=5, value=member.club_id if member else '')
        ws.cell(row=row_idx, column=6, value=ins.insurance_type.value if ins.insurance_type else '')
        ws.cell(row=row_idx, column=7, value=ins.insurance_company or '')
        ws.cell(row=row_idx, column=8, value=str(ins.coverage_amount) if ins.coverage_amount else '')
        ws.cell(row=row_idx, column=9, value=ins.status.value if ins.status else '')
        ws.cell(row=row_idx, column=10, value=ins.start_date.strftime('%d/%m/%Y') if ins.start_date else '')
        ws.cell(row=row_idx, column=11, value=ins.end_date.strftime('%d/%m/%Y') if ins.end_date else '')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"seguros_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.post("/licenses/import", response_model=ImportMembersResponse)
async def import_licenses(
    request: ImportLicensesRequest,
    create_license_use_case=Depends(get_create_license_use_case),
    update_license_use_case=Depends(get_update_license_use_case),
    license_repo=Depends(get_license_repository),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Import licenses from Excel data. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden importar licencias"
        )

    imported = 0
    updated = 0
    failed = 0
    errors = []
    is_upsert = request.mode == "upsert"

    for idx, row in enumerate(request.licenses):
        try:
            license_number = row.get('license_number') or row.get('Nº Licencia') or row.get('nº licencia') or ''
            dni = row.get('dni') or row.get('DNI') or row.get('Dni') or ''
            grade = row.get('grade') or row.get('Grado') or row.get('Grado Técnico') or row.get('grado') or ''
            technical_grade = row.get('technical_grade') or row.get('Grado Técnico') or row.get('grado_tecnico') or 'kyu'
            instructor_category = row.get('instructor_category') or row.get('Cat. Instructor') or row.get('cat_instructor') or 'none'
            age_category = row.get('age_category') or row.get('Cat. Edad') or row.get('cat_edad') or 'adulto'
            issue_date_str = row.get('issue_date') or row.get('Fecha Emisión') or row.get('fecha_emision') or None
            expiration_date_str = row.get('expiration_date') or row.get('Fecha Expiración') or row.get('fecha_expiracion') or None
            is_renewed_raw = row.get('is_renewed') or row.get('Renovada') or row.get('renovada') or None

            if not license_number:
                errors.append(f"Fila {idx + 1}: Nº de licencia es obligatorio")
                failed += 1
                continue

            # In upsert mode, find existing license by license_number and update
            if is_upsert:
                existing = await license_repo.find_by_license_number(license_number)
                if existing:
                    update_fields = {}
                    issue_date = _parse_date(issue_date_str)
                    expiration_date = _parse_date(expiration_date_str)

                    if grade:
                        update_fields['grade'] = grade
                    if technical_grade:
                        try:
                            update_fields['technical_grade'] = TechnicalGrade(technical_grade.lower())
                        except ValueError:
                            pass
                    if instructor_category:
                        try:
                            update_fields['instructor_category'] = InstructorCategory(instructor_category.lower())
                        except ValueError:
                            pass
                    if age_category:
                        try:
                            update_fields['age_category'] = AgeCategory(age_category.lower())
                        except ValueError:
                            pass
                    if issue_date:
                        update_fields['issue_date'] = issue_date
                    if expiration_date:
                        update_fields['expiration_date'] = expiration_date
                    if is_renewed_raw is not None:
                        update_fields['is_renewed'] = str(is_renewed_raw).lower() in ('sí', 'si', 'yes', 'true', '1')

                    if update_fields:
                        await update_license_use_case.execute(
                            license_id=existing.id,
                            **update_fields
                        )
                    updated += 1
                    continue

            # For create mode, grade and DNI are required
            if not grade:
                errors.append(f"Fila {idx + 1}: Grado es obligatorio")
                failed += 1
                continue

            if not dni:
                errors.append(f"Fila {idx + 1}: DNI es obligatorio para buscar el miembro")
                failed += 1
                continue

            # Lookup member by DNI
            member = await member_repo.find_by_dni(dni)
            if not member:
                errors.append(f"Fila {idx + 1}: No se encontró miembro con DNI {dni}")
                failed += 1
                continue

            issue_date = _parse_date(issue_date_str)
            expiration_date = _parse_date(expiration_date_str)

            # Validate enums
            try:
                TechnicalGrade(technical_grade.lower())
            except ValueError:
                errors.append(f"Fila {idx + 1}: Grado técnico inválido '{technical_grade}'. Use: dan, kyu")
                failed += 1
                continue

            try:
                InstructorCategory(instructor_category.lower())
            except ValueError:
                errors.append(f"Fila {idx + 1}: Categoría instructor inválida '{instructor_category}'. Use: none, fukushidoin, shidoin")
                failed += 1
                continue

            try:
                AgeCategory(age_category.lower())
            except ValueError:
                errors.append(f"Fila {idx + 1}: Categoría edad inválida '{age_category}'. Use: infantil, adulto")
                failed += 1
                continue

            await create_license_use_case.execute(
                license_number=license_number,
                member_id=member.id,
                club_id=member.club_id or '',
                grade=grade,
                license_type=technical_grade.lower(),
                issue_date=issue_date,
                expiration_date=expiration_date
            )
            imported += 1

        except Exception as e:
            failed += 1
            errors.append(f"Fila {idx + 1}: {str(e)}")

    return ImportMembersResponse(
        success=failed == 0,
        imported=imported,
        updated=updated,
        failed=failed,
        errors=errors
    )


@router.post("/insurances/import", response_model=ImportMembersResponse)
async def import_insurances(
    request: ImportInsurancesRequest,
    create_insurance_use_case=Depends(get_create_insurance_use_case),
    update_insurance_use_case=Depends(get_update_insurance_use_case),
    insurance_repo=Depends(get_insurance_repository),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Import insurances from Excel data. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden importar seguros"
        )

    imported = 0
    updated = 0
    failed = 0
    errors = []
    is_upsert = request.mode == "upsert"

    for idx, row in enumerate(request.insurances):
        try:
            policy_number = row.get('policy_number') or row.get('Nº Póliza') or row.get('nº poliza') or ''
            dni = row.get('dni') or row.get('DNI') or row.get('Dni') or ''
            ins_type = row.get('insurance_type') or row.get('Tipo Seguro') or row.get('tipo_seguro') or 'accident'
            company = row.get('insurance_company') or row.get('Compañía') or row.get('compania') or ''
            coverage_str = row.get('coverage_amount') or row.get('Cobertura') or row.get('cobertura') or None
            start_date_str = row.get('start_date') or row.get('Fecha Inicio') or row.get('fecha_inicio') or None
            end_date_str = row.get('end_date') or row.get('Fecha Fin') or row.get('fecha_fin') or None

            if not policy_number:
                errors.append(f"Fila {idx + 1}: Nº de póliza es obligatorio")
                failed += 1
                continue

            # In upsert mode, find existing insurance by policy_number and update
            if is_upsert:
                existing = await insurance_repo.find_by_policy_number(policy_number)
                if existing:
                    update_fields = {}
                    start_date = _parse_date(start_date_str)
                    end_date = _parse_date(end_date_str)

                    if ins_type:
                        ins_type_normalized = ins_type.lower().replace(' ', '_')
                        if ins_type_normalized in ('accidente', 'accident'):
                            ins_type_normalized = 'accident'
                        elif ins_type_normalized in ('rc', 'responsabilidad_civil', 'civil_liability'):
                            ins_type_normalized = 'civil_liability'
                        try:
                            update_fields['insurance_type'] = InsuranceType(ins_type_normalized)
                        except ValueError:
                            pass
                    if company:
                        update_fields['insurance_company'] = company
                    if start_date:
                        update_fields['start_date'] = start_date
                    if end_date:
                        update_fields['end_date'] = end_date
                    if coverage_str:
                        try:
                            update_fields['coverage_amount'] = float(str(coverage_str).replace(',', '.'))
                        except (ValueError, TypeError):
                            pass

                    if update_fields:
                        await update_insurance_use_case.execute(
                            insurance_id=existing.id,
                            **update_fields
                        )
                    updated += 1
                    continue

            # For create mode, company and DNI are required
            if not company:
                errors.append(f"Fila {idx + 1}: Compañía es obligatoria")
                failed += 1
                continue

            if not dni:
                errors.append(f"Fila {idx + 1}: DNI es obligatorio para buscar el miembro")
                failed += 1
                continue

            # Lookup member by DNI
            member = await member_repo.find_by_dni(dni)
            if not member:
                errors.append(f"Fila {idx + 1}: No se encontró miembro con DNI {dni}")
                failed += 1
                continue

            start_date = _parse_date(start_date_str)
            end_date = _parse_date(end_date_str)

            if not start_date or not end_date:
                errors.append(f"Fila {idx + 1}: Fechas de inicio y fin son obligatorias")
                failed += 1
                continue

            # Validate insurance type
            ins_type_normalized = ins_type.lower().replace(' ', '_')
            if ins_type_normalized in ('accidente', 'accident'):
                ins_type_normalized = 'accident'
            elif ins_type_normalized in ('rc', 'responsabilidad_civil', 'civil_liability'):
                ins_type_normalized = 'civil_liability'

            try:
                InsuranceType(ins_type_normalized)
            except ValueError:
                errors.append(f"Fila {idx + 1}: Tipo de seguro inválido '{ins_type}'. Use: accident, civil_liability")
                failed += 1
                continue

            coverage_amount = None
            if coverage_str:
                try:
                    coverage_amount = float(str(coverage_str).replace(',', '.'))
                except (ValueError, TypeError):
                    pass

            await create_insurance_use_case.execute(
                member_id=member.id,
                policy_number=policy_number,
                insurance_company=company,
                start_date=start_date,
                end_date=end_date,
                insurance_type=ins_type_normalized,
                coverage_amount=coverage_amount
            )
            imported += 1

        except Exception as e:
            failed += 1
            errors.append(f"Fila {idx + 1}: {str(e)}")

    return ImportMembersResponse(
        success=failed == 0,
        imported=imported,
        updated=updated,
        failed=failed,
        errors=errors
    )


# --- Payment type display labels ---
PAYMENT_TYPE_LABELS = {
    "licencia_kyu": "Licencia KYU",
    "licencia_kyu_infantil": "Licencia KYU Infantil",
    "licencia_dan": "Licencia DAN",
    "titulo_fukushidoin": "Título Fukushidoin",
    "titulo_shidoin": "Título Shidoin",
    "seguro_accidentes": "Seguro Accidentes",
    "seguro_rc": "Seguro RC",
    "cuota_club": "Cuota Club",
}

PAYMENT_TYPE_FROM_LABEL = {v.lower(): k for k, v in PAYMENT_TYPE_LABELS.items()}
PAYMENT_TYPE_FROM_LABEL.update({k: k for k in PAYMENT_TYPE_LABELS})


@router.get("/payments/export")
async def export_payments(
    payment_year: int = Query(..., description="Year to export payments for"),
    member_payment_repo=Depends(get_member_payment_repository),
    member_repo=Depends(get_member_repository),
    club_repo=Depends(get_club_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Export payments to Excel. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden exportar pagos"
        )

    clubs = await club_repo.find_all()

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

    ws = wb.active
    ws.title = "Pagos"

    headers = [
        "Club", "Nombre", "Apellidos", "DNI", "Tipo Pago",
        "Concepto", "Monto", "Estado", "Año"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for club in clubs:
        if not club.id or not club.is_active:
            continue

        members = await member_repo.find_by_club_id(club.id)
        member_map = {m.id: m for m in members if m.id}
        member_ids = list(member_map.keys())

        if not member_ids:
            continue

        payments = await member_payment_repo.find_by_member_ids_year(
            member_ids=member_ids,
            payment_year=payment_year
        )

        for payment in payments:
            member = member_map.get(payment.member_id)
            ws.cell(row=row_idx, column=1, value=club.name)
            ws.cell(row=row_idx, column=2, value=member.first_name if member else '')
            ws.cell(row=row_idx, column=3, value=member.last_name if member else '')
            dni = (member.dni or '') if member else ''
            ws.cell(row=row_idx, column=4, value='' if dni == 'null' else dni)
            ws.cell(row=row_idx, column=5, value=PAYMENT_TYPE_LABELS.get(payment.payment_type.value, payment.payment_type.value))
            ws.cell(row=row_idx, column=6, value=payment.concept)
            ws.cell(row=row_idx, column=7, value=payment.amount)
            ws.cell(row=row_idx, column=8, value=payment.status.value)
            ws.cell(row=row_idx, column=9, value=payment.payment_year)
            row_idx += 1

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pagos_export_{payment_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/payments/import", response_model=ImportMembersResponse)
async def import_payments(
    request: ImportPaymentsRequest,
    member_payment_repo=Depends(get_member_payment_repository),
    member_repo=Depends(get_member_repository),
    ctx: AuthContext = Depends(get_auth_context)
):
    """Import member payments from Excel data. Super admin only."""
    if not ctx.is_super_admin:
        raise HTTPException(
            status_code=403,
            detail="Solo los super administradores pueden importar pagos"
        )

    imported = 0
    updated = 0
    failed = 0
    errors = []
    is_upsert = request.mode == "upsert"

    for idx, row in enumerate(request.payments):
        try:
            dni = row.get('dni') or row.get('DNI') or row.get('Dni') or ''
            # Treat the string "null" as empty (bad data in DB)
            if isinstance(dni, str) and dni.strip().lower() == 'null':
                dni = ''
            tipo_pago = row.get('tipo_pago') or row.get('Tipo Pago') or row.get('payment_type') or ''
            year_raw = row.get('ano') or row.get('Año') or row.get('Ano') or row.get('payment_year') or ''
            monto_raw = row.get('monto') or row.get('Monto') or row.get('amount') or 0
            estado = row.get('estado') or row.get('Estado') or row.get('status') or 'completed'
            concepto = row.get('concepto') or row.get('Concepto') or row.get('concept') or ''

            if not dni:
                if is_upsert:
                    # In upsert mode, skip rows without DNI (can't match member)
                    continue
                errors.append(f"Fila {idx + 1}: DNI es obligatorio")
                failed += 1
                continue

            if not tipo_pago:
                errors.append(f"Fila {idx + 1}: Tipo de pago es obligatorio")
                failed += 1
                continue

            try:
                payment_year = int(year_raw)
            except (ValueError, TypeError):
                errors.append(f"Fila {idx + 1}: Año inválido '{year_raw}'")
                failed += 1
                continue

            try:
                amount = float(str(monto_raw).replace(',', '.'))
            except (ValueError, TypeError):
                errors.append(f"Fila {idx + 1}: Monto inválido '{monto_raw}'")
                failed += 1
                continue

            tipo_normalized = tipo_pago.lower().strip()
            payment_type_value = PAYMENT_TYPE_FROM_LABEL.get(tipo_normalized, tipo_normalized)
            try:
                payment_type = MemberPaymentType(payment_type_value)
            except ValueError:
                valid = ', '.join(PAYMENT_TYPE_LABELS.values())
                errors.append(f"Fila {idx + 1}: Tipo de pago inválido '{tipo_pago}'. Use: {valid}")
                failed += 1
                continue

            estado_normalized = estado.lower().strip()
            try:
                payment_status = MemberPaymentStatus(estado_normalized)
            except ValueError:
                errors.append(f"Fila {idx + 1}: Estado inválido '{estado}'. Use: pending, completed, refunded")
                failed += 1
                continue

            member = await member_repo.find_by_dni(dni)
            if not member:
                errors.append(f"Fila {idx + 1}: No se encontró miembro con DNI {dni}")
                failed += 1
                continue

            if is_upsert:
                existing_payments = await member_payment_repo.find_by_member_year(
                    member_id=member.id,
                    payment_year=payment_year
                )
                existing = None
                for ep in existing_payments:
                    if ep.payment_type == payment_type:
                        existing = ep
                        break

                if existing:
                    existing.amount = amount
                    existing.status = payment_status
                    if concepto:
                        existing.concept = concepto
                    existing.updated_at = datetime.utcnow()
                    await member_payment_repo.update(existing)
                    updated += 1
                    continue

            if not concepto:
                concepto = f"{PAYMENT_TYPE_LABELS.get(payment_type.value, payment_type.value)} {payment_year}"

            new_payment = MemberPayment(
                payment_id=f"import_{datetime.now().strftime('%Y%m%d%H%M%S')}_{idx}",
                member_id=member.id,
                payment_year=payment_year,
                payment_type=payment_type,
                concept=concepto,
                amount=amount,
                status=payment_status
            )
            await member_payment_repo.create(new_payment)
            imported += 1

        except Exception as e:
            failed += 1
            errors.append(f"Fila {idx + 1}: {str(e)}")

    return ImportMembersResponse(
        success=failed == 0,
        imported=imported,
        updated=updated,
        failed=failed,
        errors=errors
    )
